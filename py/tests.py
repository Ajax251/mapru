import openpyxl
import os

# ============================================================
# НАСТРОЙКИ
# ============================================================
EXCEL_FILE = 'voprosy.xlsx'       # Исходный файл Excel
OUTPUT_FILE = 'result.txt'        # Основной файл с корректными вопросами
ERRORS_FILE = 'errors.txt'        # Файл с проблемными вопросами (все ответы правильные / нет правильных)
# ============================================================


def is_cell_colored(cell) -> bool:
    """
    Возвращает True, если ячейка имеет цветную (не белую, не прозрачную) заливку.
    Учитывает RGB-цвета и тематические цвета Excel.
    """
    fill = cell.fill
    if not fill or fill.patternType != 'solid':
        return False

    color = fill.fgColor

    if color.type == 'rgb':
        rgb = color.rgb.upper() if color.rgb else ''
        # Исключаем: прозрачный (00000000), белый (FFFFFFFF), белый без альфа (FFFFFF)
        EXCLUDED_RGB = {'00000000', 'FFFFFFFF', '00FFFFFF', 'FFFFFF', ''}
        return rgb not in EXCLUDED_RGB

    elif color.type == 'theme':
        # Темы 0 (фон листа) и 1 (белый) — не цветные
        if color.theme in (0, 1):
            return False
        return True

    return False


def format_question_block(question_text: str, answers: list) -> str:
    """
    Формирует текстовый блок одного вопроса с ответами.
    answers — список кортежей (answer_text, is_correct).
    """
    lines = [f"?{question_text}"]
    for answer_text, is_correct in answers:
        prefix = '+' if is_correct else '='
        lines.append(f"{prefix}{answer_text}")
    lines.append('')  # пустая строка между вопросами
    return '\n'.join(lines)


def extract_questions_to_txt():
    print(f"Открываем файл {EXCEL_FILE}...")

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    except FileNotFoundError:
        print(f"Ошибка: Файл '{EXCEL_FILE}' не найден в папке {os.getcwd()}")
        return

    correct_blocks = []   # Вопросы с нормально определёнными ответами
    error_blocks = []     # Проблемные вопросы

    total_questions = 0
    total_errors = 0

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"Обработка листа: {sheet_name}")

        for row in sheet.iter_rows(min_row=2):
            # 3-я колонка (индекс 2) — текст вопроса
            question_cell = row[2] if len(row) > 2 else None
            if not question_cell or not question_cell.value:
                continue

            question_text = str(question_cell.value).strip().replace('\n', ' ')
            total_questions += 1

            # Собираем ответы начиная с 4-й колонки (индекс 3)
            answers = []
            for answer_cell in row[3:]:
                if not answer_cell.value:
                    continue
                answer_text = str(answer_cell.value).strip().replace('\n', ' ')
                is_correct = is_cell_colored(answer_cell)
                answers.append((answer_text, is_correct))

            if not answers:
                continue  # Вопрос без вариантов ответов — пропускаем

            correct_count = sum(1 for _, is_correct in answers if is_correct)
            total_count = len(answers)

            block = format_question_block(question_text, answers)

            # Проблемный вопрос: все ответы помечены как правильные ИЛИ ни одного правильного
            if correct_count == 0 or correct_count == total_count:
                reason = (
                    "НЕТ правильных ответов (ни одна ячейка не закрашена)"
                    if correct_count == 0
                    else f"ВСЕ {total_count} ответов помечены как правильные"
                )
                error_block = f"# ПРОБЛЕМА: {reason}\n{block}"
                error_blocks.append(error_block)
                total_errors += 1
            else:
                correct_blocks.append(block)

    # Запись основного файла
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write('\n'.join(correct_blocks))

    # Запись файла с ошибками (только если они есть)
    if error_blocks:
        with open(ERRORS_FILE, 'w', encoding='utf-8') as f:
            f.write(f"# Файл с проблемными вопросами\n")
            f.write(f"# Всего проблемных вопросов: {total_errors}\n")
            f.write(f"# Проверьте цвета ячеек в исходном Excel-файле\n\n")
            f.write('\n'.join(error_blocks))
        print(f"\n⚠️  Найдено проблемных вопросов: {total_errors}")
        print(f"   Они сохранены в: {ERRORS_FILE}")
    else:
        print(f"\n✅ Проблемных вопросов не найдено")

    print(f"\n📊 Итого обработано вопросов: {total_questions}")
    print(f"   Корректных: {total_questions - total_errors}")
    print(f"   Проблемных: {total_errors}")
    print(f"\n✅ Результат сохранён в: {OUTPUT_FILE}")


if __name__ == "__main__":
    extract_questions_to_txt()
