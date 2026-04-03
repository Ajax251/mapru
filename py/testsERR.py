import openpyxl
import os

# НАСТРОЙКИ
EXCEL_FILE = 'voprosy.xlsx'      # Имя вашего исходного файла
OUTPUT_FILE = 'result.txt'       # Имя файла, который получится на выходе

def extract_questions_to_txt():
    print(f"Открываем файл {EXCEL_FILE}...")
    try:
        # data_only=True позволяет читать значения, а не формулы
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    except FileNotFoundError:
        print(f"Ошибка: Файл '{EXCEL_FILE}' не найден в папке {os.getcwd()}")
        return

    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        # Проходимся по всем листам в документе
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"Обработка листа: {sheet_name}")
            
            # Идем по строкам, начиная со 2-й (пропуская заголовок)
            for row in sheet.iter_rows(min_row=2):
                # Проверяем, что в колонке "Вопрос" (это 3-я колонка, индекс 2) есть текст
                question_cell = row[2]
                if not question_cell.value:
                    continue  # Если вопроса нет, пропускаем строку
                
                # Убираем переносы строк внутри вопроса, чтобы он был одной строкой
                question_text = str(question_cell.value).strip().replace('\n', ' ')
                
                # Записываем вопрос
                f.write(f"?{question_text}\n")
                
                # Варианты ответов начинаются с 4-й колонки (индекс 3) и до конца
                for answer_cell in row[3:]:
                    if not answer_cell.value:
                        continue # Пропускаем пустые ячейки ответов
                        
                    answer_text = str(answer_cell.value).strip().replace('\n', ' ')
                    
                    # Проверяем заливку ячейки (цвет)
                    is_correct = False
                    
                    # Если у ячейки есть заливка (patternType) и цвет не белый/прозрачный
                    if answer_cell.fill and answer_cell.fill.patternType == 'solid':
                        color = answer_cell.fill.fgColor
                        # '00000000' и 'FFFFFFFF' обычно означают отсутствие цвета или белый фон
                        if color.type == 'rgb' and color.rgb not in ['00000000', 'FFFFFFFF', None]:
                            is_correct = True
                        elif color.theme is not None:
                            # Если использована тема оформления (тоже считается цветом)
                            is_correct = True
                    
                    # Записываем ответ с нужным знаком
                    if is_correct:
                        f.write(f"+{answer_text}\n")
                    else:
                        f.write(f"={answer_text}\n")
                
                # Пустая строка между вопросами для красоты и разделения
                f.write("\n")

    print(f"\nГотово! Результат сохранен в файл: {OUTPUT_FILE}")

if __name__ == "__main__":
    extract_questions_to_txt()